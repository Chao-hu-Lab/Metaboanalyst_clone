"""
Generate a version comparison report across control/loose/default/strict filtering groups.
Data collected from run_from_config.py console output (2026-03-23).
Corrected 2026-03-23: fixed NaN bug in multipletests (constant features after LoD imputation
caused pvalue_adj=NaN for entire ANOVA result; volcano FDR also corrected).
"""

import pandas as pd
from pathlib import Path

OUTPUT_PATH = Path(r"C:\Users\user\Desktop\MS Data process package\Metaboanalyst_clone\results\version_comparison_report.xlsx")
OUTPUT_PATH_NTU = Path(r"C:\Users\user\Desktop\NTU cancer\version_comparison_new\toolkit_DNP_MA\version_comparison_report.xlsx")

# ── Collected results ──────────────────────────────────────────────────────

VERSIONS = ["Control", "Loose", "Default", "Strict"]

data = {
    # ── Input / Feature count ──────────────────────────────────────────────
    "input_features":        [699,   277,   188,   100],
    "after_step1_features":  [674,   277,   188,   100],
    "total_cells":           [59415, 23545, 15980, 8500],
    "zero_cells":            [27662, 5196,  1192,  118],
    "final_samples":         [78,    78,    78,    78],
    "final_features":        [674,   277,   188,   100],

    # ── PCA variance explained ─────────────────────────────────────────────
    "pca_pc1_pct":           [42.6,  38.0,  33.6,  22.4],
    "pca_pc2_pct":           [36.9,  19.7,  16.7,  13.5],
    "pca_pc3_pct":           [3.5,   5.5,   6.5,   9.6],

    # ── ANOVA (3-group, FDR<0.05) ──────────────────────────────────────────
    "anova_significant":     [603,   207,   152,   79],
    "anova_total":           [674,   277,   188,   100],

    # ── OPLS-DA: Exposure vs Normal ────────────────────────────────────────
    "oplsda_ExpNor_R2Y":     [0.985, 0.979, 0.964, 0.882],
    "oplsda_ExpNor_Q2":      [0.983, 0.974, 0.953, 0.847],

    # ── OPLS-DA: Exposure vs Control ──────────────────────────────────────
    "oplsda_ExpCon_R2Y":     [0.998, 0.996, 0.992, 0.952],
    "oplsda_ExpCon_Q2":      [0.997, 0.994, 0.988, 0.936],

    # ── OPLS-DA: Normal vs Control ────────────────────────────────────────
    "oplsda_NorCon_R2Y":     [0.989, 0.991, 0.985, 0.912],
    "oplsda_NorCon_Q2":      [0.988, 0.989, 0.981, 0.889],

    # ── PLS-DA VIP max ────────────────────────────────────────────────────
    "plsda_vip_max_ExpNor":  [2.35,  3.29,  3.35,  4.00],
    "plsda_vip_max_ExpCon":  [2.34,  2.53,  2.61,  3.25],
    "plsda_vip_max_NorCon":  [2.78,  3.04,  3.09,  4.50],

    # ── Volcano: Exposure vs Normal (paired, FC≥2, FDR<0.05) ──────────────
    "volcano_ExpNor_sig":    [206, 88,  71,  25],
    "volcano_ExpNor_up":     [148, 38,  31,  11],
    "volcano_ExpNor_down":   [58,  50,  40,  14],
    "volcano_ExpNor_pairs":  [28,  28,  28,  28],

    # ── Volcano: Exposure vs Control (unpaired) ────────────────────────────
    "volcano_ExpCon_sig":    [232, 110, 81,  35],
    "volcano_ExpCon_up":     [164, 64,  39,  18],
    "volcano_ExpCon_down":   [68,  46,  42,  17],

    # ── Volcano: Normal vs Control (unpaired) ─────────────────────────────
    "volcano_NorCon_sig":    [314, 101, 79,  24],
    "volcano_NorCon_up":     [247, 54,  39,  14],
    "volcano_NorCon_down":   [67,  47,  40,  10],
}

df = pd.DataFrame(data, index=VERSIONS).T
df.index.name = "Metric"

# ── Build formatted sheets ─────────────────────────────────────────────────

# Sheet 1: Overview
overview_rows = {
    "=== Input Data ===":                   ["", "", "", ""],
    "Input features (from toolkit)":        data["input_features"],
    "Features after Step1 (miss=100% cut)": data["after_step1_features"],
    "Total cells":                          data["total_cells"],
    "Zero cells":                           data["zero_cells"],
    "Zero rate (%)":                        [f"{z/t*100:.1f}" for z, t in zip(data["zero_cells"], data["total_cells"])],
    "Final: samples × features":            [f"78 × {f}" for f in data["final_features"]],

    "=== PCA ===":                          ["", "", "", ""],
    "PC1 variance (%)":                     data["pca_pc1_pct"],
    "PC2 variance (%)":                     data["pca_pc2_pct"],
    "PC3 variance (%)":                     data["pca_pc3_pct"],
    "PC1+PC2 variance (%)":                 [round(a+b, 1) for a, b in zip(data["pca_pc1_pct"], data["pca_pc2_pct"])],

    "=== ANOVA (3-group, FDR<0.05) ===":   ["", "", "", ""],
    "Significant features":                 data["anova_significant"],
    "Total features tested":                data["anova_total"],
    "Significant rate (%)":                 [f"{s/t*100:.1f}" if t>0 else "—" for s, t in zip(data["anova_significant"], data["anova_total"])],

    "=== OPLS-DA: Exposure vs Normal ===":  ["", "", "", ""],
    "R2Y":                                  data["oplsda_ExpNor_R2Y"],
    "Q2":                                   data["oplsda_ExpNor_Q2"],
    "ΔR2Y-Q2 (overfitting gap)":            [round(r-q, 3) for r, q in zip(data["oplsda_ExpNor_R2Y"], data["oplsda_ExpNor_Q2"])],

    "=== OPLS-DA: Exposure vs Control ===": ["", "", "", ""],
    "R2Y ":                                 data["oplsda_ExpCon_R2Y"],
    "Q2 ":                                  data["oplsda_ExpCon_Q2"],
    "ΔR2Y-Q2 (overfitting gap) ":           [round(r-q, 3) for r, q in zip(data["oplsda_ExpCon_R2Y"], data["oplsda_ExpCon_Q2"])],

    "=== OPLS-DA: Normal vs Control ===":   ["", "", "", ""],
    "R2Y  ":                                data["oplsda_NorCon_R2Y"],
    "Q2  ":                                 data["oplsda_NorCon_Q2"],
    "ΔR2Y-Q2 (overfitting gap)  ":          [round(r-q, 3) for r, q in zip(data["oplsda_NorCon_R2Y"], data["oplsda_NorCon_Q2"])],

    "=== PLS-DA VIP (max) ===":             ["", "", "", ""],
    "Exposure vs Normal":                   data["plsda_vip_max_ExpNor"],
    "Exposure vs Control":                  data["plsda_vip_max_ExpCon"],
    "Normal vs Control":                    data["plsda_vip_max_NorCon"],

    "=== Volcano: Exp vs Nor (paired, FC≥2, FDR<0.05) ===": ["", "", "", ""],
    "Significant":                          data["volcano_ExpNor_sig"],
    "Up-regulated (Exposure↑)":             data["volcano_ExpNor_up"],
    "Down-regulated (Exposure↓)":           data["volcano_ExpNor_down"],
    "Matched pairs":                        data["volcano_ExpNor_pairs"],

    "=== Volcano: Exp vs Con (unpaired) ===": ["", "", "", ""],
    "Significant ":                         data["volcano_ExpCon_sig"],
    "Up-regulated (Exposure↑) ":            data["volcano_ExpCon_up"],
    "Down-regulated (Exposure↓) ":          data["volcano_ExpCon_down"],

    "=== Volcano: Nor vs Con (unpaired) ===": ["", "", "", ""],
    "Significant  ":                        data["volcano_NorCon_sig"],
    "Up-regulated (Normal↑)  ":             data["volcano_NorCon_up"],
    "Down-regulated (Normal↓)  ":           data["volcano_NorCon_down"],
}

overview_df = pd.DataFrame(overview_rows, index=VERSIONS).T
overview_df.index.name = "Metric"

# ── Write Excel ────────────────────────────────────────────────────────────

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    overview_df.to_excel(writer, sheet_name="Overview")
    df.to_excel(writer, sheet_name="Raw Data")

    # ── Apply basic formatting ─────────────────────────────────────────────
    wb = writer.book
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    SECTION_FILL  = PatternFill("solid", fgColor="D6E4F0")
    ZERO_FILL     = PatternFill("solid", fgColor="FFF2CC")   # yellow — no sig features
    SIG_FILL      = PatternFill("solid", fgColor="E2EFDA")   # green — has sig features
    HEADER_FONT   = Font(color="FFFFFF", bold=True)
    SECTION_FONT  = Font(bold=True, color="1F4E79")
    THIN          = Side(style="thin")
    BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws = wb["Overview"]

    # Header row (row 1)
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 45
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 14

    for row in ws.iter_rows(min_row=2):
        label = str(row[0].value or "")
        is_section = label.startswith("===")
        for cell in row:
            cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left")
            if is_section:
                cell.fill = SECTION_FILL
                cell.font = SECTION_FONT
            # Highlight ANOVA significance
            if "Significant features" in label and cell.column > 1:
                val = cell.value
                try:
                    cell.fill = SIG_FILL if int(val) > 0 else ZERO_FILL
                except (TypeError, ValueError):
                    pass

    # Freeze top row + first column
    ws.freeze_panes = "B2"

    # Raw Data sheet column width
    ws2 = wb["Raw Data"]
    ws2.column_dimensions["A"].width = 35
    for col in range(2, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    ws2.freeze_panes = "B2"

import shutil
shutil.copy2(OUTPUT_PATH, OUTPUT_PATH_NTU)
print(f"Report saved to: {OUTPUT_PATH}")
print(f"           also: {OUTPUT_PATH_NTU}")
