"""
Universal analysis runner — reads a YAML config and executes the full pipeline.

Usage (from project root):
    python scripts/run_from_config.py configs/step4_dnp_specnorm.yaml
    python scripts/run_from_config.py configs/step4_dnp_no_norm.yaml
    python scripts/run_from_config.py configs/dnp_pqn_normalized.yaml
"""

import argparse
import json
import os
import shutil
import sys
import warnings
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping

# Ensure project root on path (scripts/ lives one level below project root)
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

import matplotlib  # noqa: E402

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
import pandas as pd  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from core.app_config import apply_cli_overrides, dump_yaml, load_yaml_config  # noqa: E402
from core.batch_correction import build_combat_design, evaluate_combat_design  # noqa: E402
from core.feature_metadata import (  # noqa: E402
    FEATURE_MARKER_COLUMN,
    STEP4_REASON_COLUMNS,
    extract_feature_metadata,
    is_step4_ratio_column,
)
from core.input_resolver import (  # noqa: E402
    build_labels_from_sample_info,
    detect_sample_type_row_key,
    get_feature_id_column,
    infer_group_from_sample_name,
    read_input_table,
    validate_label_consistency,
    validate_sample_info_alignment,
)
from core.pipeline import MetaboAnalystPipeline  # noqa: E402
from core.sample_interface import identify_sample_columns  # noqa: E402
from core.sample_info import (
    read_sample_info_sheet,
    build_aligned_factors,
    extract_subject_ids,
)  # noqa: E402
from analysis.outlier import run_outlier_detection  # noqa: E402
from analysis.random_forest import run_random_forest  # noqa: E402
from analysis.roc import run_roc_analysis  # noqa: E402
from ms_core.analysis.pca import run_pca  # noqa: E402
from ms_core.analysis.anova import run_anova  # noqa: E402
from ms_core.analysis.univariate import volcano_analysis  # noqa: E402
from ms_core.visualization.pca_plot import plot_pca_score  # noqa: E402
from ms_core.visualization.anova_plot import plot_anova_importance, plot_feature_boxplot  # noqa: E402
from visualization.heatmap import plot_grouped_heatmap, plot_heatmap  # noqa: E402
from visualization.norm_preview import plot_norm_comparison  # noqa: E402
from visualization.oplsda_plot import plot_oplsda_splot  # noqa: E402
from visualization.outlier_plot import plot_dmodx, plot_outlier_score  # noqa: E402
from visualization.rf_plot import plot_confusion_matrix, plot_rf_importance  # noqa: E402
from visualization.roc_plot import plot_auc_ranking, plot_roc_curves  # noqa: E402

warnings.filterwarnings("ignore")

TRUE_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
FALSE_FILL = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
EVIDENCE_TIER_FILLS = {
    "Tier1_ConcordantPairwise": PatternFill(
        fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE"
    ),
    "Tier2_MultiMethod": PatternFill(
        fill_type="solid", start_color="FFD9EAF7", end_color="FFD9EAF7"
    ),
    "Tier3_SingleMethod": PatternFill(
        fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC"
    ),
    "Tier0_NoStatEvidence": PatternFill(
        fill_type="solid", start_color="FFE7E6E6", end_color="FFE7E6E6"
    ),
}
REDUNDANT_EXPORT_COLUMNS = {
    "qc_rsd_exempted",
    "qc_rsd_threshold",
    "kept_after_qc_rsd",
    "qc_rsd_pass",
    "pvalue_raw",
    "significance_pvalue",
}
REPORT_SUBDIRS = {
    "review": "00_Review_Pack",
    "qc": "01_QC_and_Preprocessing",
    "global": "02_Global_Profiling",
    "feature": "03_Feature_Selection",
    "validation": "04_Biomarker_Validation",
    "supplementary": "05_Supplementary",
}
SUMMARY_STEP4_METADATA_COLUMNS = (FEATURE_MARKER_COLUMN, *STEP4_REASON_COLUMNS)
EVIDENCE_TIER_COLUMN = "Evidence_Tier"
EVIDENCE_TIER_ORDER = {
    "Tier1_ConcordantPairwise": 0,
    "Tier2_MultiMethod": 1,
    "Tier3_SingleMethod": 2,
    "Tier0_NoStatEvidence": 3,
}


def _ensure_report_dirs(output_dir: str) -> dict[str, Path]:
    """Create publication-oriented report subdirectories under the output root."""
    root = Path(output_dir)
    report_dirs: dict[str, Path] = {}
    for key, name in REPORT_SUBDIRS.items():
        path = root / name
        path.mkdir(parents=True, exist_ok=True)
        report_dirs[key] = path
    return report_dirs


def _save_figure(
    fig: Any,
    path: Path,
    *,
    draft_mode: bool = False,
    save_pdf: bool = False,
) -> None:
    """Save a matplotlib figure as PNG, optionally adding a PDF companion."""
    dpi = 150 if draft_mode else 300
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    if save_pdf:
        fig.savefig(path.with_suffix(".pdf"), bbox_inches="tight")
    plt.close(fig)


def _copy_review_pack_figures(
    report_dirs: Mapping[str, Path],
    report_pairs: list[tuple[str, str]],
) -> int:
    """Copy curated PNG figures into the review pack directory."""
    review_dir = report_dirs["review"]
    figure_sources: list[Path] = [
        report_dirs["qc"] / "pca_score_plot.png",
        report_dirs["global"] / "heatmap_top50_grouped.png",
        report_dirs["feature"] / "anova_importance.png",
        report_dirs["supplementary"] / "plsda_score_all_groups.png",
    ]
    for group1, group2 in report_pairs:
        pair_name = f"{group1}_vs_{group2}"
        figure_sources.extend(
            [
                report_dirs["feature"] / f"oplsda_score_{pair_name}.png",
                report_dirs["feature"] / f"plsda_vip_{pair_name}.png",
                report_dirs["feature"] / f"volcano_{pair_name}.png",
            ]
        )

    copied = 0
    for source in figure_sources:
        if not source.is_file():
            continue
        copied += 1
        destination = review_dir / f"{copied:02d}_{source.name}"
        shutil.copy2(source, destination)
    return copied


def _iter_output_files(output_dir: str) -> Iterable[Path]:
    """Yield every file under the output directory in stable relative order."""
    root = Path(output_dir)
    yield from sorted(
        (path for path in root.rglob("*") if path.is_file()),
        key=lambda path: str(path.relative_to(root)),
    )


# ── Config loader ─────────────────────────────────────────


def load_config(path: str) -> dict:
    """Load and validate a YAML configuration file."""
    return load_yaml_config(path, require_required_sections=True).to_dict(
        include_runtime=True
    )


def parse_pair_config(
    raw_pairs: list,
) -> list[tuple[str, str, bool]]:
    """
    Parse volcano_pairs / oplsda_pairs config supporting both formats:
      Old: [["Exposure", "Normal"], ...]
      New: [{"groups": ["Exposure", "Normal"], "paired": true}, ...]
    Returns list of (group1, group2, paired) tuples.
    """
    result = []
    for entry in raw_pairs:
        if isinstance(entry, dict):
            groups = entry["groups"]
            paired = entry.get("paired", False)
        else:
            groups = entry
            paired = False
        result.append((groups[0], groups[1], paired))
    return result


def unique_group_pairs(
    parsed_pairs: list[tuple[str, str, bool]],
) -> list[tuple[str, str]]:
    """Return de-duplicated ordered group pairs while ignoring paired-mode metadata."""
    seen: set[tuple[str, str]] = set()
    unique_pairs: list[tuple[str, str]] = []
    for group1, group2, _paired in parsed_pairs:
        pair = (str(group1), str(group2))
        if pair in seen:
            continue
        seen.add(pair)
        unique_pairs.append(pair)
    return unique_pairs


def resolve_volcano_test_mode(volcano_cfg: Mapping[str, Any]) -> tuple[str, bool, bool]:
    """Return (test_key, equal_var, nonpar) for volcano analysis."""
    test_key = (
        str(
            volcano_cfg.get("test", volcano_cfg.get("parametric_test_default", "welch"))
        )
        .strip()
        .lower()
    )
    if test_key not in {"student", "welch", "wilcoxon"}:
        test_key = "welch"
    return test_key, test_key == "student", test_key == "wilcoxon"


def resolve_volcano_parametric_equal_var(volcano_cfg: Mapping[str, Any]) -> bool:
    """Return the equal-variance flag for unpaired parametric volcano tests."""
    _test_key, equal_var, _nonpar = resolve_volcano_test_mode(volcano_cfg)
    return equal_var


def resolve_top_vip(plsda_cfg: Mapping[str, Any]) -> int:
    """Return the configured VIP count while keeping a sane positive lower bound."""
    return max(1, int(plsda_cfg.get("top_vip", 15)))


def resolve_volcano_fc_threshold(
    volcano_cfg: Mapping[str, Any],
) -> tuple[float, float | None]:
    """Return volcano FC thresholds while supporting legacy log2-only configs."""
    fc_thresh = volcano_cfg.get("fc_thresh")
    log2_fc_thresh = volcano_cfg.get("log2_fc_thresh")
    if fc_thresh is None and log2_fc_thresh is None:
        return 2.0, None
    if fc_thresh is None:
        return float(2 ** float(log2_fc_thresh)), float(log2_fc_thresh)
    return float(fc_thresh), None if log2_fc_thresh is None else float(log2_fc_thresh)


# ── Data loaders ──────────────────────────────────────────


def _annotate_feature_table(
    df: pd.DataFrame,
    feature_metadata: pd.DataFrame | None,
    feature_column: str = "Feature",
) -> pd.DataFrame:
    if (
        feature_metadata is None
        or feature_metadata.empty
        or feature_column not in df.columns
    ):
        return df

    meta = feature_metadata.copy()
    if meta.index.name != feature_column:
        meta.index.name = feature_column
    extra_cols = [col for col in meta.columns if col not in df.columns]
    if not extra_cols:
        return df
    return df.merge(meta[extra_cols].reset_index(), on=feature_column, how="left")


def load_data(cfg: dict) -> tuple[pd.DataFrame, pd.Series, pd.DataFrame]:
    """Load Excel data and return (samples x features, labels, feature metadata)."""
    input_cfg = cfg["input"]
    input_file = input_cfg["file"]
    if not input_file:
        raise ValueError(
            "No input file specified. Use --input <path> on the command line."
        )
    fmt = input_cfg.get("format", "sample_type_row")
    is_excel_input = Path(input_file).suffix.lower() in {".xlsx", ".xls"}

    print("=" * 60)
    print(f"Loading data from: {os.path.basename(input_file)}")
    loaded_input = read_input_table(input_file)
    raw = loaded_input.table
    if loaded_input.sheet_name:
        print(f"  Selected worksheet: {loaded_input.sheet_name}")
    feature_col = get_feature_id_column(raw)
    sample_columns = [col for col in identify_sample_columns(raw) if col != feature_col]
    sample_info = read_sample_info_sheet(input_file) if is_excel_input else None
    sample_type_key = detect_sample_type_row_key(raw, feature_column=feature_col)
    if sample_type_key is not None and sample_info is None and is_excel_input:
        raise ValueError(
            f"SampleInfo sheet is required for Excel files with Sample_Type rows in '{Path(input_file).name}'."
        )

    if sample_type_key is not None:
        # Row 0 = Sample_Type labels; rows 1+ = feature values
        id_values = raw[feature_col].astype(str)
        group_rows = raw[id_values == str(sample_type_key)]
        if group_rows.empty:
            raise ValueError(
                "Sample_Type row could not be resolved from the selected worksheet."
            )
        if len(group_rows) > 1:
            raise ValueError(
                "Sample_Type row must be unique in the selected worksheet."
            )
        sample_type_row = group_rows.iloc[0]
        valid_sample_cols = [
            col for col in sample_columns if pd.notna(sample_type_row.get(col))
        ]
        feature_rows = raw[id_values != str(sample_type_key)].copy()
        feature_names = pd.Index(feature_rows[feature_col].astype(str), name="Feature")
        sample_types = sample_type_row.loc[valid_sample_cols].values
        sample_names = np.array(valid_sample_cols)

        data = feature_rows.loc[:, valid_sample_cols].values.T
        data = pd.DataFrame(data, columns=feature_names, index=sample_names)
        data = data.apply(pd.to_numeric, errors="coerce")
        labels = pd.Series(sample_types, index=sample_names, name="Group")
        if sample_info is not None:
            labels = validate_label_consistency(
                data.index,
                labels,
                sample_info,
                observed_label_name="Worksheet Sample_Type",
            )
        feature_metadata = extract_feature_metadata(
            feature_rows.reset_index(drop=True), feature_names
        )

    elif fmt == "plain":
        # All rows are features; groups inferred from column names
        feature_names = pd.Index(raw[feature_col].astype(str), name="Feature")
        sample_names = np.array(sample_columns)

        data = raw.loc[:, sample_columns].values.T
        data = pd.DataFrame(data, columns=feature_names, index=sample_names)
        data = data.apply(pd.to_numeric, errors="coerce")
        feature_metadata = extract_feature_metadata(
            raw.reset_index(drop=True), feature_names
        )
        if sample_info is not None:
            labels = build_labels_from_sample_info(
                data.index, sample_info, label_name="Group"
            )
        elif input_cfg.get("plain_label_mode") == "column_names":
            labels = pd.Series(sample_names, index=sample_names, name="Group")
        else:
            labels = pd.Series(
                [infer_group_from_sample_name(n) for n in sample_names],
                index=sample_names,
                name="Group",
            )

        # Drop excluded samples
        exclude_mask = labels == "__EXCLUDE__"
        if exclude_mask.any():
            excluded = list(sample_names[exclude_mask])
            print(f"  Excluding {len(excluded)} samples: {excluded}")
            data = data.loc[~exclude_mask]
            labels = labels.loc[~exclude_mask]
    else:
        raise ValueError(
            f"Unknown input format: '{fmt}'. Use 'sample_type_row' or 'plain'."
        )

    if sample_info is not None:
        validate_sample_info_alignment(data.index, sample_info)

    print(f"  Samples: {data.shape[0]}, Features: {data.shape[1]}")
    print(f"  Groups: {dict(labels.value_counts())}")
    print(f"  Zeros: {(data == 0).sum().sum()} / {data.size}")
    print(f"  Missing: {data.isna().sum().sum()} / {data.size}")
    marker_count = int(feature_metadata[FEATURE_MARKER_COLUMN].sum())
    print(f"  Presence/absence markers: {marker_count}")
    if feature_metadata.attrs.get("step4_metadata_detected", False):
        ratio_columns = [
            str(column)
            for column in feature_metadata.columns
            if is_step4_ratio_column(column)
        ]
        reason_columns = [
            column for column in STEP4_REASON_COLUMNS if column in feature_metadata.columns
        ]
        print("  Step4 metadata detected")
        print(f"  Step4 ratio columns: {len(ratio_columns)}")
        print(
            "  Step4 reason columns: "
            f"{', '.join(reason_columns) if reason_columns else 'none'}"
        )

    return data, labels, feature_metadata


def _finalize_analysis_matrix(
    df: pd.DataFrame,
    labels: pd.Series,
    included_groups: list[str],
) -> tuple[pd.DataFrame, pd.Series]:
    """Apply the same sample/feature filtering used before downstream analyses."""
    qc_mask = labels.astype(str).str.contains("QC", case=False, na=False)
    if qc_mask.any():
        df = df.loc[~qc_mask]
        labels = labels.loc[~qc_mask]

    if included_groups:
        group_mask = labels.isin(included_groups)
        df = df.loc[group_mask]
        labels = labels.loc[group_mask]

    feat_std = df.std()
    zero_var_feats = feat_std[feat_std == 0].index.tolist()
    if zero_var_feats:
        df = df.drop(columns=zero_var_feats)

    return df, labels


def compose_output_suffix(
    base_suffix: str = "",
    *,
    include_timestamp: bool = True,
    timestamp: str | None = None,
) -> str:
    """Build the final output suffix, appending a timestamp by default."""
    base = (base_suffix or "").strip()
    if base and not base.startswith("_"):
        base = f"_{base}"

    if not include_timestamp:
        return base

    ts = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{ts}" if base else f"_{ts}"


def _select_heatmap_matrix(
    df: pd.DataFrame,
    anova_df: pd.DataFrame,
    *,
    max_features: int,
) -> pd.DataFrame:
    """Choose a stable subset for publication heatmap export."""
    feature_cap = max(1, min(int(max_features), 50))
    ranked = anova_df.sort_values("pvalue_adj").copy()
    if "significant" in ranked.columns:
        significant_mask = ranked["significant"].fillna(False).astype(bool)
        significant = ranked.loc[significant_mask]
    else:
        significant = ranked.iloc[0:0]
    feature_rows = significant if not significant.empty else ranked
    feature_names = [
        feat
        for feat in feature_rows["Feature"].astype(str).tolist()
        if feat in df.columns
    ]
    selected = feature_names[:feature_cap]
    if selected:
        return df.loc[:, selected].copy()
    return df.copy()


def _summary_evidence_family(sheet_name: str) -> str | None:
    """Return the summary evidence family represented by an Excel sheet name."""
    normalized = sheet_name.strip().lower()
    if normalized.startswith("anova"):
        return "anova"
    if normalized.startswith("vip_"):
        return "vip"
    if normalized.startswith("volcano_"):
        return "volcano"
    return None


def _summary_pair_key(sheet_name: str, family: str | None) -> str | None:
    """Return a normalized pair key for pairwise evidence sheets."""
    if family not in {"vip", "volcano"}:
        return None
    prefix = f"{family}_"
    normalized = sheet_name.strip().lower()
    if not normalized.startswith(prefix):
        return None
    pair = normalized[len(prefix) :].strip()
    return pair or None


def _top15_features_for_evidence(sheet_name: str, df: pd.DataFrame) -> set[str]:
    """Return the feature set eligible for top-15 pair concordance checks."""
    family = _summary_evidence_family(sheet_name)
    if family not in {"vip", "volcano"} or "Feature" not in df.columns:
        return set()

    if family == "vip":
        if "Rank" in df.columns:
            ranks = pd.to_numeric(df["Rank"], errors="coerce")
            top_df = df.loc[ranks <= 15]
        else:
            top_df = df.head(15)
        return set(top_df["Feature"].astype(str))

    sortable = df.copy()
    if "pvalue_adj" in sortable.columns:
        sortable["_pvalue_adj_sort"] = pd.to_numeric(
            sortable["pvalue_adj"], errors="coerce"
        )
    else:
        sortable["_pvalue_adj_sort"] = np.nan
    top_df = sortable.sort_values(
        ["_pvalue_adj_sort", "Feature"], ascending=[True, True], na_position="last"
    ).head(15)
    return set(top_df["Feature"].astype(str))


def _classify_evidence_tier(record: Mapping[str, Any] | None) -> str:
    """Classify a feature-level cross-method evidence record."""
    if not record:
        return "Tier0_NoStatEvidence"

    families = set(record.get("families", set()))
    vip_top15_pairs = set(record.get("vip_top15_pairs", set()))
    volcano_top15_pairs = set(record.get("volcano_top15_pairs", set()))
    vip_pairs = set(record.get("vip_pairs", set()))
    volcano_pairs = set(record.get("volcano_pairs", set()))

    if vip_top15_pairs & volcano_top15_pairs:
        return "Tier1_ConcordantPairwise"
    if "anova" in families and (vip_pairs & volcano_pairs):
        return "Tier1_ConcordantPairwise"
    if len(families) >= 2:
        return "Tier2_MultiMethod"
    if len(families) == 1:
        return "Tier3_SingleMethod"
    return "Tier0_NoStatEvidence"


# ── Significant features Excel export ────────────────────


def _export_significant_features_excel(
    sheets: dict,
    output_dir: str,
    top_n: int | None = None,
):
    """
    Write a consolidated Excel workbook with significant features from all analyses.

    Sheets collected during the pipeline are written as-is, then a 'Summary'
    sheet is generated that cross-references feature appearances across methods.
    """
    if not sheets:
        print("  No significant feature data collected — skipping Excel export.")
        return

    excel_path = os.path.join(output_dir, "significant_features_summary.xlsx")
    summary_csv_path = os.path.join(output_dir, "Summary.csv")
    summary_sheet_filter = {
        sheet_name: df
        for sheet_name, df in sheets.items()
        if "oplsda" not in sheet_name.lower()
    }
    feature_metadata_by_feature: dict[str, dict[str, Any]] = {}
    summary_metadata_columns: list[str] = []
    excel_detail_metadata_columns: list[str] = []
    evidence_records: dict[str, dict[str, set[str]]] = {}

    def is_excel_detail_metadata_column(column: str) -> bool:
        return column in STEP4_REASON_COLUMNS or is_step4_ratio_column(column)

    def apply_step4_excel_outline(worksheet) -> None:
        headers = [cell.value for cell in worksheet[1]]
        detail_indexes = [
            index
            for index, header in enumerate(headers, start=1)
            if isinstance(header, str) and is_excel_detail_metadata_column(header)
        ]
        if not detail_indexes:
            return

        worksheet.sheet_properties.outlinePr.summaryRight = True
        max_column = len(headers)
        ranges: list[tuple[int, int]] = []
        start = previous = detail_indexes[0]
        for index in detail_indexes[1:]:
            if index == previous + 1:
                previous = index
                continue
            ranges.append((start, previous))
            start = previous = index
        ranges.append((start, previous))

        for start, end in ranges:
            for index in range(start, end + 1):
                dimension = worksheet.column_dimensions[get_column_letter(index)]
                dimension.hidden = True
                dimension.outlineLevel = 1
            collapsed_index = min(end + 1, max_column)
            worksheet.column_dimensions[get_column_letter(collapsed_index)].collapsed = True

    def apply_evidence_tier_style(worksheet) -> None:
        headers = [cell.value for cell in worksheet[1]]
        if EVIDENCE_TIER_COLUMN not in headers:
            return
        tier_col = headers.index(EVIDENCE_TIER_COLUMN) + 1
        worksheet.column_dimensions[get_column_letter(tier_col)].width = 28
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=tier_col)
            fill = EVIDENCE_TIER_FILLS.get(str(cell.value))
            if fill is not None:
                cell.fill = fill

    def build_summary_excel_df(summary_df: pd.DataFrame) -> pd.DataFrame:
        if summary_df.empty or not excel_detail_metadata_columns:
            return summary_df

        excel_df = summary_df.copy()
        for column in excel_detail_metadata_columns:
            if column in excel_df.columns:
                continue
            excel_df[column] = excel_df["Feature"].map(
                lambda feat: feature_metadata_by_feature.get(feat, {}).get(column, "")
            )

        ordered_columns: list[str] = []
        inserted_details = False
        for column in excel_df.columns:
            if column in excel_detail_metadata_columns:
                continue
            ordered_columns.append(column)
            if column == FEATURE_MARKER_COLUMN:
                ordered_columns.extend(
                    detail
                    for detail in excel_detail_metadata_columns
                    if detail in excel_df.columns
                )
                inserted_details = True
        if not inserted_details:
            ordered_columns.extend(
                detail
                for detail in excel_detail_metadata_columns
                if detail in excel_df.columns and detail not in ordered_columns
            )
        return excel_df.loc[:, ordered_columns]

    def passes_threshold(sheet_name: str, row: pd.Series) -> bool:
        name = sheet_name.lower()
        p_adj = pd.to_numeric(row.get("pvalue_adj"), errors="coerce")
        vip = pd.to_numeric(row.get("VIP"), errors="coerce")
        log2fc = pd.to_numeric(row.get("log2FC"), errors="coerce")
        importance = pd.to_numeric(row.get("Importance"), errors="coerce")
        loading = pd.to_numeric(row.get("Loading"), errors="coerce")

        if "anova" in name:
            return pd.notna(p_adj) and p_adj < 0.05
        if "vip" in name:
            return pd.notna(vip) and vip > 1.0
        if "volcano" in name:
            return (
                pd.notna(p_adj)
                and p_adj < 0.05
                and pd.notna(log2fc)
                and abs(log2fc) >= 1.0
            )
        if "oplsda" in name:
            if pd.notna(vip):
                return vip > 1.0
            if pd.notna(importance):
                return abs(importance) > 0.05
            if pd.notna(loading):
                return abs(loading) > 0.05
            return False
        return False

    def prepare_export_sheet(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
        sheet_df = df.copy()
        if "Feature" not in sheet_df.columns:
            return sheet_df

        if "significant" not in sheet_df.columns:
            sheet_df["significant"] = sheet_df.apply(
                lambda row: passes_threshold(sheet_name, row),
                axis=1,
            )

        drop_cols = [col for col in REDUNDANT_EXPORT_COLUMNS if col in sheet_df.columns]
        if drop_cols:
            sheet_df = sheet_df.drop(columns=drop_cols)

        preferred_order = [
            "Rank",
            "Feature",
            FEATURE_MARKER_COLUMN,
            "imputation_method",
            "VIP",
            "Importance",
            "Loading",
            "log2FC",
            "pvalue",
            "pvalue_adj",
            "neg_log10p",
            "statistic",
            "qc_detect_ratio",
            "qc_rsd",
            "significant",
        ]
        existing = [col for col in preferred_order if col in sheet_df.columns]
        remainder = [col for col in sheet_df.columns if col not in existing]
        return sheet_df.loc[:, existing + remainder]

    def score_text(sheet_name: str, row: pd.Series) -> str:
        name = sheet_name.lower()
        if "vip" in name and pd.notna(row.get("VIP")):
            return f"VIP={float(row['VIP']):.2f}"
        if (
            "volcano" in name
            and pd.notna(row.get("pvalue_adj"))
            and pd.notna(row.get("log2FC"))
        ):
            return f"p_adj={float(row['pvalue_adj']):.2e};log2FC={float(row['log2FC']):.2f}"
        if "anova" in name and pd.notna(row.get("pvalue_adj")):
            return f"p_adj={float(row['pvalue_adj']):.2e}"
        if "oplsda" in name:
            if pd.notna(row.get("Importance")):
                return f"|p|={float(row['Importance']):.4f}"
            if pd.notna(row.get("Loading")):
                return f"loading={float(row['Loading']):.4f}"
        return "Pass"

    # Build cross-reference summary with complete feature coverage.
    feature_counts = {}  # feature -> {sheet_name: score_text}
    all_features: set[str] = set()
    prepared_sheets = {
        sheet_name: prepare_export_sheet(sheet_name, df)
        for sheet_name, df in sheets.items()
    }

    for sheet_name, df in {
        name: prepared_sheets[name] for name in summary_sheet_filter
    }.items():
        if "Feature" not in df.columns:
            continue
        evidence_family = _summary_evidence_family(sheet_name)
        evidence_pair_key = _summary_pair_key(sheet_name, evidence_family)
        metadata_columns = [
            column for column in SUMMARY_STEP4_METADATA_COLUMNS if column in df.columns
        ]
        excel_metadata_columns = [
            column for column in df.columns if is_excel_detail_metadata_column(column)
        ]
        sheet_df = df if top_n in (None, 0) else df.head(top_n)
        top15_features = _top15_features_for_evidence(sheet_name, sheet_df)
        for idx, row in sheet_df.iterrows():
            feat = row["Feature"]
            all_features.add(feat)
            if feat not in feature_metadata_by_feature:
                feature_metadata_by_feature[feat] = {}
            for column in excel_metadata_columns:
                if column not in excel_detail_metadata_columns:
                    excel_detail_metadata_columns.append(column)
            for column in metadata_columns:
                if column not in summary_metadata_columns:
                    summary_metadata_columns.append(column)
            for column in excel_metadata_columns:
                if column not in summary_metadata_columns:
                    summary_metadata_columns.append(column)
            for column in list(dict.fromkeys(metadata_columns + excel_metadata_columns)):
                value = row.get(column)
                if column in feature_metadata_by_feature[feat] and pd.isna(value):
                    continue
                feature_metadata_by_feature[feat][column] = value
            passed = passes_threshold(sheet_name, row)
            if passed and evidence_family is not None:
                record = evidence_records.setdefault(
                    feat,
                    {
                        "families": set(),
                        "vip_pairs": set(),
                        "volcano_pairs": set(),
                        "vip_top15_pairs": set(),
                        "volcano_top15_pairs": set(),
                    },
                )
                record["families"].add(evidence_family)
                if evidence_pair_key is not None and evidence_family in {
                    "vip",
                    "volcano",
                }:
                    record[f"{evidence_family}_pairs"].add(evidence_pair_key)
                    if str(feat) in top15_features:
                        record[f"{evidence_family}_top15_pairs"].add(evidence_pair_key)
            if not passed:
                continue
            if feat not in feature_counts:
                feature_counts[feat] = {}
            feature_counts[feat][sheet_name] = score_text(sheet_name, row)

    # Build summary DataFrame
    all_sheet_names = [
        s
        for s in summary_sheet_filter.keys()
        if "Feature" in summary_sheet_filter[s].columns
    ]
    summary_rows = []
    for feat in sorted(all_features):
        appearances = feature_counts.get(feat, {})
        metadata_values = feature_metadata_by_feature.get(feat, {})
        row = {
            "Feature": feat,
            EVIDENCE_TIER_COLUMN: _classify_evidence_tier(
                evidence_records.get(feat)
            ),
            "Passed_in_N_analyses": len(appearances),
        }
        for column in summary_metadata_columns:
            default_value = False if column == FEATURE_MARKER_COLUMN else ""
            row[column] = metadata_values.get(column, default_value)
        for sn in all_sheet_names:
            row[sn] = appearances.get(sn, "")
        summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows)
    if not summary_df.empty:
        summary_df["_Evidence_Tier_Rank"] = summary_df[EVIDENCE_TIER_COLUMN].map(
            EVIDENCE_TIER_ORDER
        )
        summary_df = summary_df.sort_values(
            ["_Evidence_Tier_Rank", "Passed_in_N_analyses", "Feature"],
            ascending=[True, False, True],
        ).reset_index(drop=True)
        summary_df = summary_df.drop(columns=["_Evidence_Tier_Rank"])
        summary_df.insert(0, "Rank", range(1, len(summary_df) + 1))
        summary_df.to_csv(summary_csv_path, index=False)
    summary_excel_df = build_summary_excel_df(summary_df)

    # Write all sheets
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Summary first
        if not summary_excel_df.empty:
            summary_excel_df.to_excel(writer, sheet_name="Summary", index=False)

        # Individual analysis sheets (top_n rows each)
        for sheet_name, df in prepared_sheets.items():
            # Truncate sheet name to Excel's 31-char limit
            safe_name = sheet_name[:31]
            sheet_df = df if top_n in (None, 0) else df.head(top_n)
            sheet_df.to_excel(writer, sheet_name=safe_name, index=False)

        workbook = writer.book
        if not summary_excel_df.empty:
            summary_worksheet = workbook["Summary"]
            apply_evidence_tier_style(summary_worksheet)
            apply_step4_excel_outline(summary_worksheet)
        for sheet_name, df in prepared_sheets.items():
            sheet_df = df if top_n in (None, 0) else df.head(top_n)
            safe_name = sheet_name[:31]
            apply_step4_excel_outline(workbook[safe_name])
            if "significant" not in sheet_df.columns:
                continue
            worksheet = workbook[safe_name]
            significant_col = sheet_df.columns.get_loc("significant") + 1
            for row_idx, value in enumerate(sheet_df["significant"], start=2):
                cell = worksheet.cell(row=row_idx, column=significant_col)
                cell.fill = TRUE_FILL if bool(value) else FALSE_FILL

    print(f"  Saved {excel_path}")
    if not summary_df.empty:
        print(f"  Saved {summary_csv_path}")
    print(f"  Sheets: {', '.join(['Summary'] + list(sheets.keys()))}")
    scope_text = "all rows" if top_n in (None, 0) else f"top {top_n} each"
    print(
        f"  Summary: {len(summary_df)} unique features across "
        f"{len(all_sheet_names)} analyses ({scope_text})"
    )


# ── Main analysis ─────────────────────────────────────────


def run_analysis(cfg: dict) -> dict[str, str]:
    """Execute the full analysis pipeline from a config dict."""

    data, labels, feature_metadata = load_data(cfg)
    pipe_cfg = cfg["pipeline"]
    combat_cfg = cfg.get("combat", {})
    analysis_cfg = cfg["analysis"]

    # ── Resolve output directory ──────────────────────────
    results_root = os.path.join(_PROJECT_ROOT, "results")
    input_stem = os.path.splitext(os.path.basename(cfg["input"]["file"]))[0]
    base_suffix = cfg["output"].get("suffix", "")
    auto_timestamp = bool(cfg["output"].get("auto_timestamp", True))
    resolved_suffix = compose_output_suffix(
        base_suffix, include_timestamp=auto_timestamp
    )
    cfg["output"]["base_suffix"] = base_suffix
    cfg["output"]["suffix"] = resolved_suffix
    output_dir = os.path.join(results_root, input_stem + resolved_suffix)
    os.makedirs(output_dir, exist_ok=True)
    report_dirs = _ensure_report_dirs(output_dir)

    output_cfg = cfg.get("output", {})
    draft_mode = bool(output_cfg.get("draft_mode", False))
    save_pdf = bool(output_cfg.get("save_pdf", False))
    if not draft_mode:
        from visualization.theme import apply_publication_export_style

        apply_publication_export_style()

    # ── Build SpecNorm factors if needed ──────────────────
    factors = None
    factor_source = None
    batch_labels = None
    combat_covariates = None
    combat_source = None
    sample_info = None
    combat_par_prior = bool(combat_cfg.get("par_prior", True))
    combat_mean_only = bool(combat_cfg.get("mean_only", False))
    combat_ref_batch = combat_cfg.get("ref_batch")
    requires_sample_info = (
        pipe_cfg.get("row_norm") == "SpecNorm"
        or pipe_cfg.get("batch_correction") == "ComBat"
    )
    if requires_sample_info:
        sample_info = read_sample_info_sheet(cfg["input"]["file"])

    if pipe_cfg.get("row_norm") == "SpecNorm":
        spec_cfg = cfg.get("spec_norm")
        if not spec_cfg or "factor_column" not in spec_cfg:
            raise ValueError(
                "row_norm='SpecNorm' requires a 'spec_norm' section with 'factor_column'."
            )
        print("\n" + "=" * 60)
        print("Loading SampleInfo for concentration correction...")
        if sample_info is None:
            raise RuntimeError("SampleInfo sheet not found in Excel file!")

        factor_col = spec_cfg["factor_column"]
        factor_source = spec_cfg.get("factor_source", factor_col)
        factors, meta = build_aligned_factors(sample_info, data.index, factor_col)
        print(f"  Factor column: {factor_col}")
        print(f"  Aligned: {meta['n_samples']} samples")
        print(f"  Range: {meta['min_factor']:.2f} ~ {meta['max_factor']:.2f}")
        print(f"  Fuzzy matches: {meta['n_fuzzy_matches']}")
        print(f"  QC skipped (factor=1.0): {meta['n_qc_skipped']}")

    if pipe_cfg.get("batch_correction") == "ComBat":
        if sample_info is None:
            raise RuntimeError("SampleInfo sheet with Batch column is required for ComBat.")
        covariate_mode = str(combat_cfg.get("covariate_mode", "labels")).strip().lower()
        aligned_labels = (
            labels if isinstance(labels, pd.Series) else pd.Series(labels, index=data.index)
        )
        combat_labels = aligned_labels if covariate_mode == "labels" else None
        covariate_columns = (
            combat_cfg.get("sample_info_covariates", [])
            if covariate_mode == "sample_info"
            else None
        )
        if covariate_mode == "sample_info" and not covariate_columns:
            raise ValueError(
                "combat.covariate_mode='sample_info' requires combat.sample_info_covariates."
            )
        batch_labels, combat_covariates, combat_meta = build_combat_design(
            data.index,
            sample_info,
            labels=combat_labels,
            covariate_columns=covariate_columns,
        )
        validation_covariates = combat_covariates
        if validation_covariates is None and aligned_labels.nunique(dropna=True) > 1:
            validation_covariates = aligned_labels.astype("string").rename(
                "Current labels"
            ).to_frame()
        combat_validation = evaluate_combat_design(batch_labels, validation_covariates)
        if combat_validation["blocking_errors"]:
            raise ValueError("\n".join(combat_validation["blocking_errors"]))
        combat_source = f"{combat_meta['batch_source']} ({covariate_mode})"
        print(f"  ComBat batch source: {combat_source}")
        print(f"  Batches: {combat_meta['batch_counts']}")
        if combat_meta["covariate_columns"]:
            print(f"  ComBat covariates: {', '.join(combat_meta['covariate_columns'])}")
        else:
            print("  ComBat covariates: None")
        for warning in combat_validation["warnings"]:
            print(f"  ComBat warning: {warning}")
        print(f"  ComBat mean_only: {combat_mean_only}")
        print(f"  ComBat par_prior: {combat_par_prior}")
        print(f"  ComBat ref_batch: {combat_ref_batch if combat_ref_batch else 'None'}")

    # ── Run pipeline ──────────────────────────────────────
    print("\n" + "=" * 60)
    print("Running preprocessing pipeline...")
    pipeline = MetaboAnalystPipeline(data, labels, feature_metadata=feature_metadata)

    processed = pipeline.run_pipeline(
        missing_thresh=pipe_cfg.get("missing_thresh", 0.5),
        impute_method=pipe_cfg.get("impute_method", "min"),
        filter_method=pipe_cfg.get("filter_method", "iqr"),
        filter_cutoff=pipe_cfg.get("filter_cutoff"),
        qc_rsd_enabled=pipe_cfg.get("qc_rsd_enabled", False),
        qc_rsd_threshold=pipe_cfg.get("qc_rsd_threshold", 0.20),
        row_norm=pipe_cfg.get("row_norm", "None"),
        transform=pipe_cfg.get("transform", "None"),
        batch_correction=pipe_cfg.get("batch_correction", "None"),
        scaling=pipe_cfg.get("scaling", "None"),
        factors=factors,
        factor_source=factor_source,
        batch_labels=batch_labels,
        combat_covariates=combat_covariates,
        combat_par_prior=combat_par_prior,
        combat_mean_only=combat_mean_only,
        combat_ref_batch=combat_ref_batch,
        combat_source=combat_source,
    )

    print("\nPipeline log:")
    for line in pipeline.log:
        print(f"  {line}")
    print(
        f"\nProcessed data: {processed.shape[0]} samples x {processed.shape[1]} features"
    )

    included_groups = cfg["groups"].get("include", [])
    final_labels = (
        pipeline.processed_labels if pipeline.processed_labels is not None else labels
    )
    final_feature_metadata = (
        pipeline.processed_feature_metadata
        if pipeline.processed_feature_metadata is not None
        else feature_metadata
    ).copy()

    qc_mask = final_labels.astype(str).str.contains("QC", case=False, na=False)
    if qc_mask.any():
        print(f"  Removing {qc_mask.sum()} remaining QC samples...")

    if included_groups:
        group_mask = final_labels.isin(included_groups)
        if not group_mask.all():
            excluded_count = int((~group_mask).sum())
            print(f"  Removing {excluded_count} samples not in {included_groups}")

    feat_std = processed.std()
    zero_var_feats = feat_std[feat_std == 0].index.tolist()
    if zero_var_feats:
        print(
            f"  Dropping {len(zero_var_feats)} zero-variance features (constant after QC removal)"
        )

    processed, final_labels = _finalize_analysis_matrix(
        processed, final_labels, included_groups
    )
    final_feature_metadata = final_feature_metadata.reindex(processed.columns).copy()
    volcano_matrix, _ = _finalize_analysis_matrix(
        pipeline.steps["batch_corrected"].copy(),
        (
            pipeline.processed_labels
            if pipeline.processed_labels is not None
            else labels
        ).copy(),
        included_groups,
    )
    anova_matrix = volcano_matrix.copy()
    volcano_fc_matrix, _ = _finalize_analysis_matrix(
        pipeline.steps["row_normed"].copy(),
        (
            pipeline.processed_labels
            if pipeline.processed_labels is not None
            else labels
        ).copy(),
        included_groups,
    )
    pre_norm_matrix, _ = _finalize_analysis_matrix(
        pipeline.steps["filtered"].copy(),
        (
            pipeline.processed_labels
            if pipeline.processed_labels is not None
            else labels
        ).copy(),
        included_groups,
    )
    pre_norm_matrix = pre_norm_matrix.reindex(
        index=processed.index, columns=processed.columns
    )

    raw_volcano_pairs = parse_pair_config(cfg["groups"].get("volcano_pairs", []))
    raw_oplsda_pairs = parse_pair_config(cfg["groups"].get("oplsda_pairs", []))
    report_pairs = unique_group_pairs(raw_volcano_pairs or raw_oplsda_pairs)

    print(f"  Final: {processed.shape[0]} samples x {processed.shape[1]} features")
    print(f"  Groups: {dict(final_labels.value_counts())}")

    # Collectors for the summary Excel export
    _excel_sheets = {}  # sheet_name -> DataFrame
    paired_resolution_audit_rows: list[dict[str, Any]] = []

    # Save processed data
    processed.to_csv(os.path.join(report_dirs["qc"], "processed_data.csv"))
    final_labels.to_csv(os.path.join(report_dirs["qc"], "sample_labels.csv"))
    final_feature_metadata.to_csv(
        os.path.join(report_dirs["qc"], "feature_metadata.csv"),
        index_label="Feature",
    )
    qc_rsd_audit = pipeline.step_feature_metadata.get("qc_rsd")
    if qc_rsd_audit is not None and not qc_rsd_audit.empty:
        qc_rsd_audit.to_csv(
            os.path.join(report_dirs["qc"], "qc_rsd_audit.csv"),
            index_label="Feature",
        )

    # Save a copy of the config used
    config_copy_path = os.path.join(report_dirs["qc"], "config_used.yaml")
    with open(config_copy_path, "w", encoding="utf-8") as f:
        f.write(dump_yaml(cfg, include_runtime=False))

    log_path = os.path.join(report_dirs["qc"], "pipeline_log.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        for line in pipeline.log:
            f.write(line + "\n")
    print(f"  Saved to {output_dir}")

    # ── QC / preprocessing report ────────────────────────
    print("\n" + "=" * 60)
    print("Generating QC and preprocessing report...")

    fig = plt.figure(figsize=(12, 8))
    plot_norm_comparison(pre_norm_matrix, processed, final_labels, fig=fig)
    _save_figure(
        fig,
        report_dirs["qc"] / "normalization_comparison.png",
        draft_mode=draft_mode,
        save_pdf=save_pdf,
    )
    print(f"  Saved {REPORT_SUBDIRS['qc']}\\normalization_comparison.png")

    # ── PCA ───────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Running PCA...")
    n_comp = analysis_cfg["pca"]["n_components"]
    pca_result = run_pca(processed, final_labels, n_components=n_comp)
    evr = pca_result.explained_variance_ratio
    print(
        f"  Variance explained (PC1-{min(n_comp, 5)}): "
        f"{[f'{v * 100:.1f}%' for v in evr[:5]]}"
    )

    fig = plt.figure(figsize=(10, 8))
    plot_pca_score(pca_result, pc_x=0, pc_y=1, fig=fig)
    _save_figure(
        fig,
        report_dirs["qc"] / "pca_score_plot.png",
        draft_mode=draft_mode,
        save_pdf=save_pdf,
    )
    print(f"  Saved {REPORT_SUBDIRS['qc']}\\pca_score_plot.png")

    # ── ANOVA ─────────────────────────────────────────────
    print("\n" + "=" * 60)
    anova_cfg = analysis_cfg["anova"]
    group_list = sorted(final_labels.unique())
    print(f"Running ANOVA ({' vs '.join(group_list)})...")
    anova_result = run_anova(
        anova_matrix,
        final_labels,
        p_thresh=anova_cfg["p_thresh"],
        nonpar=anova_cfg["nonpar"],
        use_fdr=anova_cfg["use_fdr"],
        posthoc=anova_cfg["posthoc"],
    )
    print(
        f"  Significant (FDR < {anova_cfg['p_thresh']}): "
        f"{anova_result.n_significant} / {len(anova_result.result_df)}"
    )

    anova_result.result_df = _annotate_feature_table(
        anova_result.result_df, final_feature_metadata
    )
    anova_result.result_df.to_csv(
        os.path.join(report_dirs["feature"], "anova_results.csv"), index=False
    )

    # Collect full ANOVA table for Excel export and summary scoring.
    _anova_sig = anova_result.result_df.sort_values("pvalue_adj")[
        [
            "Feature",
            "pvalue",
            "pvalue_adj",
            "neg_log10p",
            "significant",
            FEATURE_MARKER_COLUMN,
        ]
    ].copy()
    _excel_sheets["ANOVA_All"] = _anova_sig

    fig = plt.figure(figsize=(10, 8))
    plot_anova_importance(anova_result, top_n=25, fig=fig)
    _save_figure(
        fig,
        report_dirs["feature"] / "anova_importance.png",
        draft_mode=draft_mode,
        save_pdf=save_pdf,
    )

    anova_ranked = anova_result.result_df.sort_values("pvalue_adj").copy()
    anova_feature_pool = anova_ranked["Feature"].tolist()
    sig_feature_pool = anova_ranked.loc[anova_ranked["significant"], "Feature"].tolist()
    heatmap_cfg = analysis_cfg.get("heatmap", {})
    heatmap_df = _select_heatmap_matrix(
        processed,
        anova_ranked,
        max_features=int(heatmap_cfg.get("max_features", 50)),
    )
    if not heatmap_df.empty:
        heatmap_fig = plot_heatmap(
            heatmap_df,
            final_labels.reindex(heatmap_df.index),
            method=str(heatmap_cfg.get("method", "ward")),
            metric=str(heatmap_cfg.get("metric", "euclidean")),
            scale=heatmap_cfg.get("scale", "row"),
            max_features=min(int(heatmap_cfg.get("max_features", 50)), 50),
            top_by=str(heatmap_cfg.get("top_by", "var")),
        )
        _save_figure(
            heatmap_fig,
            report_dirs["global"] / "heatmap_top50.png",
            draft_mode=draft_mode,
            save_pdf=save_pdf,
        )
        print(f"  Saved {REPORT_SUBDIRS['global']}\\heatmap_top50.png")

        grouped_heatmap_fig = plot_grouped_heatmap(
            heatmap_df,
            final_labels.reindex(heatmap_df.index),
            group_order=included_groups,
            scale=heatmap_cfg.get("scale", "row"),
            max_features=None,
            top_by=str(heatmap_cfg.get("top_by", "var")),
        )
        _save_figure(
            grouped_heatmap_fig,
            report_dirs["global"] / "heatmap_top50_grouped.png",
            draft_mode=draft_mode,
            save_pdf=save_pdf,
        )
        print(f"  Saved {REPORT_SUBDIRS['global']}\\heatmap_top50_grouped.png")

    anova_pairs = report_pairs
    saved_anova_boxplots = 0
    anova_boxplot_top_n = 10
    for g1, g2 in anova_pairs:
        pair_mask = final_labels.isin([g1, g2])
        pair_data = anova_matrix.loc[pair_mask]
        pair_labels = final_labels.loc[pair_mask]
        if pair_data.empty or pair_labels.nunique() < 2:
            continue

        candidate_features = (
            sig_feature_pool if sig_feature_pool else anova_feature_pool
        )
        pair_means = pair_data.groupby(pair_labels).mean(numeric_only=True)
        if g1 not in pair_means.index or g2 not in pair_means.index:
            continue

        ranked_features = sorted(
            candidate_features,
            key=lambda feat: abs(pair_means.loc[g1, feat] - pair_means.loc[g2, feat]),
            reverse=True,
        )[:anova_boxplot_top_n]

        for i, feat in enumerate(ranked_features, start=1):
            fig = plt.figure(figsize=(7, 5))
            plot_feature_boxplot(pair_data, pair_labels, feat, fig=fig)
            _save_figure(
                fig,
                report_dirs["supplementary"] / f"anova_boxplot_{g1}_vs_{g2}_top{i}.png",
                draft_mode=draft_mode,
                save_pdf=save_pdf,
            )
            saved_anova_boxplots += 1
    print(f"  Saved {REPORT_SUBDIRS['feature']}\\anova_importance.png")
    print(f"  Saved {saved_anova_boxplots} supplementary ANOVA boxplots")

    # ── PLS-DA + VIP (pairwise, 2-group) ─────────────────────
    plsda_cfg = analysis_cfg.get("plsda", {})
    plsda_pairs = raw_oplsda_pairs
    if plsda_cfg and plsda_pairs:
        print("\n" + "=" * 60)
        print("Running pairwise PLS-DA + VIP...")
        try:
            from ms_core.analysis.plsda import run_plsda
            from ms_core.visualization.plsda_plot import plot_plsda_score
            from ms_core.visualization.vip_plot import plot_vip
        except ImportError as e:
            print(f"  Import error: {e}")
            plsda_pairs = []

        n_comp = plsda_cfg.get("n_components", 2)
        top_vip = resolve_top_vip(plsda_cfg)

        try:
            all_plsda_result = run_plsda(processed, final_labels, n_components=n_comp)
            fig = plt.figure(figsize=(8, 6))
            plot_plsda_score(all_plsda_result, fig=fig)
            _save_figure(
                fig,
                report_dirs["supplementary"] / "plsda_score_all_groups.png",
                draft_mode=draft_mode,
                save_pdf=save_pdf,
            )
            print(
                f"  Saved {REPORT_SUBDIRS['supplementary']}\\plsda_score_all_groups.png"
            )
        except Exception as e:
            print(f"  All-group PLS-DA error: {e}")

        for g1, g2, _paired in plsda_pairs:
            print(f"  {g1} vs {g2}...")
            try:
                pair_mask = final_labels.isin([g1, g2])
                pair_data = processed.loc[pair_mask]
                pair_labels = final_labels.loc[pair_mask]

                plsda_result = run_plsda(pair_data, pair_labels, n_components=n_comp)
                print(
                    f"    VIP range: {plsda_result.vips.min():.2f} - "
                    f"{plsda_result.vips.max():.2f}"
                )

                # Collect VIP scores for Excel export
                vip_df = _annotate_feature_table(
                    plsda_result.get_vip_df(), final_feature_metadata
                )
                vip_df.insert(0, "Rank", range(1, len(vip_df) + 1))
                _excel_sheets[f"VIP_{g1}_vs_{g2}"] = vip_df

                fig = plt.figure(figsize=(10, max(6, top_vip * 0.32)))
                plot_vip(
                    plsda_result,
                    top_n=top_vip,
                    data=pair_data,
                    labels=pair_labels,
                    fig=fig,
                )
                _save_figure(
                    fig,
                    report_dirs["feature"] / f"plsda_vip_{g1}_vs_{g2}.png",
                    draft_mode=draft_mode,
                    save_pdf=save_pdf,
                )
                print(
                    f"    Saved {REPORT_SUBDIRS['feature']}\\plsda_vip_{g1}_vs_{g2}.png"
                )
            except Exception as e:
                print(f"    Error: {e}")

    # ── OPLS-DA (pairwise, 2-group only) ──────────────────
    raw_oplsda_pairs = cfg["groups"].get("oplsda_pairs", [])
    oplsda_parsed = parse_pair_config(raw_oplsda_pairs)
    if oplsda_parsed:
        print("\n" + "=" * 60)
        print("Running pairwise OPLS-DA...")
        try:
            from ms_core.analysis.oplsda import run_oplsda
            from ms_core.visualization.oplsda_plot import plot_oplsda_score
        except ImportError as e:
            print(f"  Import error: {e}")
            oplsda_parsed = []

        for g1, g2, _paired in oplsda_parsed:
            print(f"  {g1} vs {g2}...")
            try:
                pair_mask = final_labels.isin([g1, g2])
                pair_data = processed.loc[pair_mask]
                pair_labels = final_labels.loc[pair_mask]

                oplsda_result = run_oplsda(pair_data, pair_labels)
                print(
                    f"    R2Y={oplsda_result.r2y:.3f}, Q2={oplsda_result.q2:.3f}, "
                    f"backend={oplsda_result.backend}"
                )

                # Collect OPLS-DA loadings for Excel export
                imp_df = _annotate_feature_table(
                    oplsda_result.get_importance_df(), final_feature_metadata
                )
                if not imp_df.empty:
                    imp_df.insert(0, "Rank", range(1, len(imp_df) + 1))
                    _excel_sheets[f"OPLSDA_{g1}_vs_{g2}"] = imp_df

                # Score plot
                fig = plt.figure(figsize=(8, 6))
                plot_oplsda_score(oplsda_result, fig=fig)
                _save_figure(
                    fig,
                    report_dirs["feature"] / f"oplsda_score_{g1}_vs_{g2}.png",
                    draft_mode=draft_mode,
                    save_pdf=save_pdf,
                )

                fig = plt.figure(figsize=(8, 6))
                plot_oplsda_splot(oplsda_result, top_n=10, fig=fig)
                _save_figure(
                    fig,
                    report_dirs["feature"] / f"oplsda_splot_{g1}_vs_{g2}.png",
                    draft_mode=draft_mode,
                    save_pdf=save_pdf,
                )

                print(
                    f"    Saved {REPORT_SUBDIRS['feature']}\\oplsda_score_{g1}_vs_{g2}.png"
                )
                print(
                    f"    Saved {REPORT_SUBDIRS['feature']}\\oplsda_splot_{g1}_vs_{g2}.png"
                )
            except Exception as e:
                print(f"    Error: {e}")

    # ── Volcano ───────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Running pairwise volcano analyses...")
    from visualization.volcano_plot import plot_volcano

    vol_cfg = analysis_cfg["volcano"]
    volcano_pairs = raw_volcano_pairs
    volcano_test_key, volcano_equal_var, volcano_nonpar = resolve_volcano_test_mode(
        vol_cfg
    )
    volcano_fc_thresh, volcano_log2_fc_thresh = resolve_volcano_fc_threshold(vol_cfg)
    paired_resolution_cfg = cfg["groups"].get("paired_resolution")

    # Extract subject IDs if any pair is paired
    pair_id_pattern = cfg["groups"].get("pair_id_pattern", r"BC\d+")
    has_any_paired = any(p for _, _, p in volcano_pairs)
    subject_ids = None
    if has_any_paired:
        subject_ids = extract_subject_ids(processed.index, pattern=pair_id_pattern)
        n_with_id = (subject_ids != "").sum()
        print(
            f"  Subject IDs extracted: {n_with_id}/{len(subject_ids)} "
            f"(pattern: {pair_id_pattern})"
        )

    for g1, g2, is_paired in volcano_pairs:
        test_label = "paired" if is_paired else "unpaired"
        print(f"  {g1} vs {g2} ({test_label})...")
        try:
            if volcano_nonpar:
                print(f"    Configured volcano test: {volcano_test_key}")
            vresult = volcano_analysis(
                volcano_matrix,
                final_labels,
                group1=g1,
                group2=g2,
                fc_thresh=volcano_fc_thresh,
                log2_fc_thresh=volcano_log2_fc_thresh,
                p_thresh=vol_cfg["p_thresh"],
                equal_var=volcano_equal_var,
                nonpar=volcano_nonpar,
                use_fdr=vol_cfg["use_fdr"],
                paired=is_paired,
                pair_ids=subject_ids if is_paired else None,
                fc_df=volcano_fc_matrix,
                pair_resolution=paired_resolution_cfg if is_paired else None,
            )
            pair_info = f", Pairs: {vresult.n_pairs}" if is_paired else ""
            print(f"    Method: {vresult.test_label}")
            print(
                f"    Significant: {vresult.n_significant} "
                f"(Up: {vresult.n_up}, Down: {vresult.n_down}{pair_info})"
            )
            if is_paired and vresult.resolution_overrides_applied:
                print(
                    f"    Paired overrides applied: {len(vresult.resolution_overrides_applied)}"
                )
                for record in vresult.resolution_overrides_applied:
                    paired_resolution_audit_rows.append(
                        {
                            "pair": f"{g1}_vs_{g2}",
                            "audit_kind": "override",
                            "group": record.get("group", ""),
                            "subject_id": record.get("subject_id", ""),
                            "selected_sample": record.get("selected_sample", ""),
                            "message": "",
                            "candidates": " | ".join(record.get("candidates", [])),
                        }
                    )
            if is_paired and vresult.resolution_warnings:
                print(
                    f"    Paired resolution warnings: {len(vresult.resolution_warnings)}"
                )
                for warning in vresult.resolution_warnings:
                    print(f"      - {warning}")
                    paired_resolution_audit_rows.append(
                        {
                            "pair": f"{g1}_vs_{g2}",
                            "audit_kind": "warning",
                            "group": "",
                            "subject_id": "",
                            "selected_sample": "",
                            "message": warning,
                            "candidates": "",
                        }
                    )
            vresult.result_df = _annotate_feature_table(
                vresult.result_df, final_feature_metadata
            )
            vresult.result_df.to_csv(
                os.path.join(report_dirs["feature"], f"volcano_{g1}_vs_{g2}.csv"),
                index=False,
            )

            # Collect volcano significant features for Excel export
            vol_sig = _annotate_feature_table(
                vresult.significant.copy(), final_feature_metadata
            )
            if not vol_sig.empty:
                vol_sig = vol_sig.sort_values("pvalue_adj").head(50)
                _excel_sheets[f"Volcano_{g1}_vs_{g2}"] = vol_sig
            fig = plt.figure(figsize=(10, 8))
            plot_volcano(vresult, fig=fig)
            _save_figure(
                fig,
                report_dirs["feature"] / f"volcano_{g1}_vs_{g2}.png",
                draft_mode=draft_mode,
                save_pdf=save_pdf,
            )
            print(f"    Saved {REPORT_SUBDIRS['feature']}\\volcano_{g1}_vs_{g2}.png")
        except Exception as e:
            print(f"    Error: {e}")

    # ── ROC / biomarker validation ───────────────────────
    print("\n" + "=" * 60)
    print("Running biomarker validation plots...")
    roc_cfg = analysis_cfg.get("roc", {})
    roc_top_n = int(roc_cfg.get("top_n", 10))
    roc_multi_feature = bool(roc_cfg.get("multi_feature", True))
    roc_cv_folds = int(roc_cfg.get("cv_folds", 5))
    for g1, g2 in report_pairs:
        print(f"  ROC {g1} vs {g2}...")
        try:
            pair_mask = final_labels.isin([g1, g2])
            pair_data = processed.loc[pair_mask]
            pair_labels = final_labels.loc[pair_mask]
            if pair_data.empty or pair_labels.nunique() < 2:
                print("    Skipped (not enough pairwise samples)")
                continue

            roc_result = run_roc_analysis(
                pair_data,
                pair_labels,
                group1=g1,
                group2=g2,
                top_n=roc_top_n,
                multi_feature=roc_multi_feature,
                cv_folds=roc_cv_folds,
            )
            roc_result.summary_df.to_csv(
                os.path.join(report_dirs["validation"], f"roc_{g1}_vs_{g2}.csv"),
                index=False,
            )

            fig = plt.figure(figsize=(8, 6))
            plot_roc_curves(roc_result, fig=fig)
            _save_figure(
                fig,
                report_dirs["validation"] / f"roc_{g1}_vs_{g2}.png",
                draft_mode=draft_mode,
                save_pdf=save_pdf,
            )

            fig = plt.figure(figsize=(8, 6))
            plot_auc_ranking(roc_result, fig=fig)
            _save_figure(
                fig,
                report_dirs["validation"] / f"auc_ranking_{g1}_vs_{g2}.png",
                draft_mode=draft_mode,
                save_pdf=save_pdf,
            )

            print(f"    Saved {REPORT_SUBDIRS['validation']}\\roc_{g1}_vs_{g2}.png")
            print(
                f"    Saved {REPORT_SUBDIRS['validation']}\\auc_ranking_{g1}_vs_{g2}.png"
            )
        except Exception as e:
            print(f"    Error: {e}")

    # ── Supplementary: outlier detection ─────────────────
    print("\n" + "=" * 60)
    print("Generating supplementary outlier plots...")
    outlier_cfg = analysis_cfg.get("outlier", {})
    try:
        outlier_result = run_outlier_detection(
            processed,
            n_components=int(outlier_cfg.get("n_components", 2)),
            alpha=float(outlier_cfg.get("alpha", 0.05)),
        )
        outlier_result.get_outlier_df().to_csv(
            os.path.join(report_dirs["supplementary"], "outlier_results.csv"),
            index=False,
        )

        fig = plt.figure(figsize=(8, 6))
        plot_outlier_score(outlier_result, labels=final_labels, fig=fig)
        _save_figure(
            fig,
            report_dirs["supplementary"] / "outlier_t2.png",
            draft_mode=draft_mode,
            save_pdf=save_pdf,
        )

        fig = plt.figure(figsize=(8, 6))
        plot_dmodx(outlier_result, labels=final_labels, fig=fig)
        _save_figure(
            fig,
            report_dirs["supplementary"] / "outlier_dmodx.png",
            draft_mode=draft_mode,
            save_pdf=save_pdf,
        )
        print(f"  Saved {REPORT_SUBDIRS['supplementary']}\\outlier_t2.png")
        print(f"  Saved {REPORT_SUBDIRS['supplementary']}\\outlier_dmodx.png")
    except Exception as e:
        print(f"  Outlier export error: {e}")

    # ── Supplementary: random forest ─────────────────────
    print("\n" + "=" * 60)
    print("Generating supplementary Random Forest plots...")
    rf_cfg = analysis_cfg.get("random_forest", {})
    rf_trees = int(rf_cfg.get("n_trees", 500))
    rf_cv_folds = int(rf_cfg.get("cv_folds", 5))
    rf_top_n = int(rf_cfg.get("top_n", 25))

    # Pass 1: collect RF results and find global vmax for confusion matrices
    rf_results: list[tuple[str, str, Any]] = []
    global_cm_vmax = 0
    for g1, g2 in report_pairs:
        print(f"  RF {g1} vs {g2}...")
        try:
            pair_mask = final_labels.isin([g1, g2])
            pair_data = processed.loc[pair_mask]
            pair_labels = final_labels.loc[pair_mask]
            if pair_data.empty or pair_labels.nunique() < 2:
                print("    Skipped (not enough pairwise samples)")
                continue

            rf_result = run_random_forest(
                pair_data,
                pair_labels,
                n_trees=rf_trees,
                cv_folds=rf_cv_folds,
                top_n=rf_top_n,
            )
            rf_results.append((g1, g2, rf_result))
            global_cm_vmax = max(global_cm_vmax, int(rf_result.confusion_mat.max()))
        except Exception as e:
            print(f"    Error: {e}")

    # Pass 2: render figures with unified color scale
    for g1, g2, rf_result in rf_results:
        rf_result.feature_importance.to_csv(
            os.path.join(
                report_dirs["supplementary"], f"rf_importance_{g1}_vs_{g2}.csv"
            ),
            index=False,
        )

        fig = plt.figure(figsize=(8, 6))
        plot_rf_importance(rf_result, top_n=rf_top_n, fig=fig)
        _save_figure(
            fig,
            report_dirs["supplementary"] / f"rf_importance_{g1}_vs_{g2}.png",
            draft_mode=draft_mode,
            save_pdf=save_pdf,
        )

        fig = plt.figure(figsize=(6, 5))
        plot_confusion_matrix(
            rf_result, fig=fig, vmax=global_cm_vmax if len(rf_results) > 1 else None
        )
        _save_figure(
            fig,
            report_dirs["supplementary"] / f"rf_confusion_matrix_{g1}_vs_{g2}.png",
            draft_mode=draft_mode,
            save_pdf=save_pdf,
        )

        print(
            f"    Saved {REPORT_SUBDIRS['supplementary']}\\rf_importance_{g1}_vs_{g2}.png"
        )
        print(
            f"    Saved {REPORT_SUBDIRS['supplementary']}\\rf_confusion_matrix_{g1}_vs_{g2}.png"
        )

    if paired_resolution_audit_rows:
        audit_path = os.path.join(
            report_dirs["supplementary"], "paired_resolution_audit.csv"
        )
        pd.DataFrame(paired_resolution_audit_rows).to_csv(audit_path, index=False)
        print(f"  Saved {audit_path}")

    # ── Export significant features summary Excel ─────────
    print("\n" + "=" * 60)
    print("Exporting significant features summary...")
    export_top_n = cfg.get("output", {}).get("export_top_n")
    try:
        _export_significant_features_excel(
            _excel_sheets,
            output_dir,
            top_n=export_top_n,
        )
    except Exception as e:
        print(f"  Excel export error: {e}")

    print("\n" + "=" * 60)
    print("Building review pack...")
    copied_review_figures = _copy_review_pack_figures(report_dirs, report_pairs)
    print(
        f"  Saved {copied_review_figures} curated PNG(s) to "
        f"{REPORT_SUBDIRS['review']}"
    )

    # ── Summary ───────────────────────────────────────────
    print("\n" + "=" * 60)
    print("ANALYSIS COMPLETE")
    print(f"All outputs saved to: {output_dir}")
    print("\nGenerated files:")
    output_root = Path(output_dir)
    for path in _iter_output_files(output_dir):
        rel = str(path.relative_to(output_root))
        size = path.stat().st_size
        print(f"  {rel:70s} ({size / 1024:.0f} KB)")
    return {"output_dir": output_dir}


# ── CLI entry point ───────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Run MetaboAnalyst pipeline from a YAML config file.",
        epilog="Example: python run_from_config.py configs/step4_dnp_specnorm.yaml",
    )
    parser.add_argument(
        "config",
        help="Path to YAML configuration file (e.g., configs/step4_dnp_specnorm.yaml)",
    )
    parser.add_argument(
        "--input",
        "-i",
        help="Override the input file path from config (e.g., path/to/data.xlsx)",
        default=None,
    )
    parser.add_argument(
        "--suffix",
        "-s",
        help="Append a suffix to the output folder name (e.g., _control, _strict)",
        default=None,
    )
    args = parser.parse_args()

    if not os.path.isfile(args.config):
        print(f"Error: Config file not found: {args.config}")
        sys.exit(1)

    app_config = load_yaml_config(args.config, require_required_sections=True)
    app_config = apply_cli_overrides(
        app_config,
        input_file=args.input,
        suffix=args.suffix,
    )
    cfg = app_config.to_dict(include_runtime=True)
    if args.input:
        print(f"Input overridden: {args.input}")
    print(f"Loaded config: {args.config}")
    result = run_analysis(cfg)
    print(f"__RESULT_JSON__:{json.dumps(str(result.get('output_dir', '')))}")


if __name__ == "__main__":
    main()
